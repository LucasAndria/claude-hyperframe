#!/usr/bin/env python3
"""Per-video pipeline: videos/<CODE>/ (script + config) -> HeyGen avatar -> Hyperframe -> ffmpeg -> final MP4.

Everything for one video lives in its own folder:

    videos/<CODE>/
        config.json      per-video settings (optional -- defaults below)
        script.xlsx      the validated script
        source-video/    (alt) a pre-recorded base clip instead of HeyGen
        public/          HyperFrames overlay composition (index.html)
        output/          all rendered results for this video

Usage:
    python main.py EMO14_VID01          # or: python main.py videos/EMO14_VID01
"""

import json
import logging
import shutil
import subprocess
import sys
from pathlib import Path

# ---------------------------------------------------------------------------
# Setup
# ---------------------------------------------------------------------------
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("avatar-pipeline")

ROOT = Path(__file__).parent
VIDEOS_DIR = ROOT / "videos"

DEFAULT_CONFIG = {
    "source": "avatar",
    "script_file": "script.xlsx",
    "output_dir": "output",
    "orientation": "horizontal",
    "heygen": {
        # Emy + French female voice Audrey (La Petite Creche defaults)
        "avatar_id": "75255783465d403696e639a249a0b9c0",
        "voice_id": "7459d7aa599b4f97908600896a0e7ef4",
    },
    "source_video": {"file": "source-video/input.mp4"},
    "hyperframe": {"fps": 25, "composition_dir": "public"},
    "motion": {"config": "motion/motion.config.json"},
}


class PipelineError(Exception):
    """Any expected, handled failure in the pipeline."""


def list_videos() -> list[str]:
    if not VIDEOS_DIR.exists():
        return []
    return sorted(p.name for p in VIDEOS_DIR.iterdir() if p.is_dir())


def resolve_video_dir(arg: str) -> Path:
    """Accept 'EMO14_VID01', 'videos/EMO14_VID01', or an absolute path."""
    for candidate in (Path(arg), ROOT / arg, VIDEOS_DIR / arg):
        if candidate.is_dir() and candidate.resolve() != ROOT.resolve():
            return candidate.resolve()
    available = "\n  ".join(list_videos()) or "(none yet -- create one with: python new_video.py <CODE>)"
    raise PipelineError(f"Video folder not found: {arg}\nAvailable videos:\n  {available}")


def _merge(base: dict, override: dict) -> dict:
    out = dict(base)
    for k, v in override.items():
        out[k] = _merge(base[k], v) if isinstance(v, dict) and isinstance(base.get(k), dict) else v
    return out


def load_config(video_dir: Path) -> dict:
    """Per-video config.json merged over defaults; missing file just means defaults."""
    cfg_path = video_dir / "config.json"
    if not cfg_path.exists():
        log.info("No config.json in %s -- using defaults.", video_dir.name)
        return dict(DEFAULT_CONFIG)
    return _merge(DEFAULT_CONFIG, json.loads(cfg_path.read_text(encoding="utf-8")))


def resolve_path(video_dir: Path, value: str) -> Path:
    """Resolve a config path: absolute as-is, else relative to the video folder,
    falling back to the repo root (for shared assets like motion/)."""
    p = Path(value)
    if p.is_absolute():
        return p
    local = video_dir / p
    return local if local.exists() else (ROOT / p if (ROOT / p).exists() else local)


# orientation -> HeyGen aspectRatio
ASPECT = {"vertical": "9:16", "horizontal": "16:9"}


def aspect_ratio(cfg: dict) -> str:
    """Map config 'orientation' (vertical|horizontal) to a HeyGen aspect ratio."""
    orient = str(cfg.get("orientation", "vertical")).strip().lower()
    if orient not in ASPECT:
        raise PipelineError(f"Invalid orientation '{orient}'. Use 'vertical' or 'horizontal'.")
    return ASPECT[orient]


def read_script(path: Path) -> str:
    """Read script text from .txt/.md/.csv or .xlsx (extensible)."""
    if not path.exists():
        raise PipelineError(f"Script file not found: {path}")

    suffix = path.suffix.lower()
    if suffix in (".txt", ".md", ".csv"):
        text = path.read_text(encoding="utf-8")
    elif suffix == ".xlsx":
        text = _read_xlsx(path)
    else:
        raise PipelineError(f"Unsupported script format: {suffix} (use .txt, .md, .csv, or .xlsx)")

    text = text.strip()
    if not text:
        raise PipelineError(f"Script file is empty: {path}")
    log.info("Read script (%d chars) from %s", len(text), path)
    return text


def _read_xlsx(path: Path) -> str:
    """Join all non-empty cells (row by row) into one script string."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise PipelineError("Reading .xlsx needs openpyxl. Install it: pip install openpyxl") from e
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        lines = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if cells:
                    lines.append(" ".join(cells))
        wb.close()
    except Exception as e:
        raise PipelineError(f"Failed to read xlsx: {e}") from e
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Step 1: HeyGen avatar video
# ---------------------------------------------------------------------------
def generate_heygen_video(script: str, cfg: dict, out_path: Path) -> Path:
    """Generate an avatar video from the script text using the HeyGen MCP.

    Uses the HeyGen MCP (not the REST API) — auth is handled by the MCP server,
    so no API key is stored here.

    TODO: Wire this to the HeyGen MCP. Have Claude run this script and call the
    MCP tool `create_video_from_avatar` with:
      - script      = the text passed in
      - avatar_id   = avatar_id (below)
      - voice_id    = voice_id (below; omit it entirely if empty so HeyGen uses
                      the avatar's default voice)
      - aspectRatio = aspect (below; from config "orientation")
    then download the finished video to out_path.
    On any failure, raise PipelineError.
    """
    avatar_id = cfg.get("heygen", {}).get("avatar_id", "")
    voice_id = cfg.get("heygen", {}).get("voice_id", "").strip()
    aspect = aspect_ratio(cfg)
    if not avatar_id:
        raise PipelineError("Missing heygen.avatar_id in config.")
    log.info("Generating HeyGen avatar video (avatar=%s, voice=%s, aspect=%s)...",
             avatar_id, voice_id or "<avatar default>", aspect)
    try:
        # TODO: real generation — download the MCP-generated video to out_path.
        # (No stub write here: writing a placeholder would clobber an existing real download.)
        raise PipelineError("HeyGen generation not implemented yet (placeholder).")
    except PipelineError:
        raise
    except Exception as e:  # network, auth, timeout, etc.
        raise PipelineError(f"HeyGen generation failed: {e}") from e
    log.info("HeyGen video saved to %s", out_path)
    return out_path


# ---------------------------------------------------------------------------
# Step 1 (alt): use a pre-recorded source video instead of a HeyGen avatar
# ---------------------------------------------------------------------------
def use_source_video(video_dir: Path, cfg: dict) -> Path:
    """Resolve the base video from the video folder's source-video/ subfolder.

    Used when config "source" == "video": the pipeline skips HeyGen generation
    and uses this file as the base video. It keeps its own audio (the silent
    Hyperframe render is muxed over it, then this audio is muxed back in).
    """
    src = str(cfg.get("source_video", {}).get("file", "")).strip()
    if not src:
        raise PipelineError('source is "video" but source_video.file is not set in config.')
    path = resolve_path(video_dir, src)
    if not path.exists() or path.stat().st_size == 0:
        raise PipelineError(f"Source video not found or empty: {path}")
    log.info("Using source video: %s", path)
    return path


# ---------------------------------------------------------------------------
# Step 2: Hyperframe motion / enhancement
# ---------------------------------------------------------------------------
def stage_motion_assets(public_dir: Path, video_dir: Path, cfg: dict) -> None:
    """Copy the shared reference-explainer motion library + global style into a composition.

    Writes public/vendor/reference-motion.js and public/vendor/motion-style.js so any
    composition can `<script src="vendor/reference-motion.js">` and use window.RefMotion
    with the project-wide motionStyle (the single place to tune intensity globally).
    """
    vendor = public_dir / "vendor"
    vendor.mkdir(parents=True, exist_ok=True)
    lib = ROOT / "motion" / "reference-motion.js"
    if lib.exists():
        shutil.copyfile(lib, vendor / "reference-motion.js")
    cfg_path = resolve_path(video_dir, cfg.get("motion", {}).get("config", "motion/motion.config.json"))
    motion_style = {}
    if cfg_path.exists():
        motion_style = json.loads(cfg_path.read_text(encoding="utf-8")).get("motionStyle", {})
    (vendor / "motion-style.js").write_text(
        "window.MOTION_STYLE = " + json.dumps(motion_style) + ";\n", encoding="utf-8")


def enhance_with_hyperframe(in_path: Path, video_dir: Path, cfg: dict, out_path: Path) -> Path:
    """Add motion / typography overlays by rendering a Hyperframes composition.

    The composition itself (timed keyword/typography cards over the base video)
    is authored once via the `graphic-overlays` skill into the video folder's
    `public/` directory. This function just renders it to video.

    Convention: cfg["hyperframe"]["composition_dir"] points to the composition
    dir relative to the video folder (default: "public"). If no composition
    exists yet, the avatar video passes through unchanged so the pipeline still
    completes.

    NOTE: `hyperframes render` produces a SILENT video — the narration audio is
    muxed back in by process_with_ffmpeg().
    """
    log.info("Enhancing with Hyperframe...")
    if not in_path.exists() or in_path.stat().st_size == 0:
        raise PipelineError(f"Hyperframe input missing or empty: {in_path}")

    comp = resolve_path(video_dir, cfg.get("hyperframe", {}).get("composition_dir") or "public")
    if not (comp / "index.html").exists():
        log.warning("No Hyperframe composition at %s — passing video through unchanged.", comp)
        shutil.copyfile(in_path, out_path)
        return out_path

    stage_motion_assets(comp, video_dir, cfg)  # ensure the motion library + global style are present
    fps = str(cfg.get("hyperframe", {}).get("fps", 25))
    cmd = ["npx", "hyperframes", "render", str(comp), "-o", str(out_path.resolve()), "--fps", fps]
    try:
        subprocess.run(cmd, check=True, capture_output=True, text=True, shell=(sys.platform == "win32"))
    except subprocess.CalledProcessError as e:
        raise PipelineError(f"Hyperframe render failed:\n{e.stderr}") from e
    except FileNotFoundError as e:
        raise PipelineError("npx/hyperframes not found. Install Node.js and run `npx hyperframes doctor`.") from e
    log.info("Hyperframe output saved to %s", out_path)
    return out_path


# ---------------------------------------------------------------------------
# Step 3: ffmpeg processing (ensure clean MP4)
# ---------------------------------------------------------------------------
def process_with_ffmpeg(in_path: Path, out_path: Path, audio_from: Path | None = None) -> Path:
    """Produce a clean MP4 (H.264/AAC, faststart).

    If audio_from is given, take video from in_path and audio from that file —
    the Hyperframe render is silent, so we mux the avatar narration back in here.
    Extend for trim/resize/merge as needed.
    """
    log.info("Processing with ffmpeg -> %s", out_path)
    if shutil.which("ffmpeg") is None:
        raise PipelineError("ffmpeg not found on PATH. Install it: https://ffmpeg.org/download.html")
    if not in_path.exists() or in_path.stat().st_size == 0:
        raise PipelineError(f"ffmpeg input missing or empty: {in_path}")

    cmd = ["ffmpeg", "-y", "-i", str(in_path)]
    if audio_from and audio_from.exists():
        cmd += ["-i", str(audio_from), "-map", "0:v:0", "-map", "1:a:0", "-shortest"]
    cmd += ["-c:v", "libx264", "-pix_fmt", "yuv420p", "-c:a", "aac", "-movflags", "+faststart", str(out_path)]
    try:
        subprocess.run(cmd, check=True, capture_output=True, text=True)
    except subprocess.CalledProcessError as e:
        raise PipelineError(f"ffmpeg error:\n{e.stderr}") from e
    log.info("Final video saved to %s", out_path)
    return out_path


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------
def main() -> int:
    try:
        if len(sys.argv) < 2:
            available = "\n  ".join(list_videos()) or "(none yet)"
            raise PipelineError(
                "Usage: python main.py <video>  (e.g. python main.py EMO14_VID01)\n"
                f"Available videos:\n  {available}")

        video_dir = resolve_video_dir(sys.argv[1])
        name = video_dir.name
        cfg = load_config(video_dir)
        out_dir = video_dir / cfg.get("output_dir", "output")
        out_dir.mkdir(exist_ok=True)
        log.info("Video project: %s", video_dir)

        source = str(cfg.get("source", "avatar")).strip().lower()
        if source == "video":
            # Use a pre-recorded clip from the video folder — no HeyGen generation.
            raw = use_source_video(video_dir, cfg)
        elif source == "avatar":
            # Generate the base video from the script with a HeyGen avatar.
            script_path = resolve_path(video_dir, cfg.get("script_file", "script.xlsx"))
            script = read_script(script_path)
            raw = generate_heygen_video(script, cfg, out_dir / f"{name}_1_heygen.mp4")
        else:
            raise PipelineError(f"Invalid source '{source}'. Use 'avatar' or 'video'.")

        enhanced = enhance_with_hyperframe(raw, video_dir, cfg, out_dir / f"{name}_2_hyperframe.mp4")
        # Hyperframe render is silent — mux the audio from the base video
        # (HeyGen narration, or the source video's own audio).
        final = process_with_ffmpeg(enhanced, out_dir / f"{name}.mp4", audio_from=raw)

        log.info("Done. Final video: %s", final)
        return 0
    except PipelineError as e:
        log.error("%s", e)
        return 1
    except Exception as e:  # unexpected
        log.exception("Unexpected error: %s", e)
        return 2


if __name__ == "__main__":
    sys.exit(main())
